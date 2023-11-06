import random
import re
import os
from os import listdir
import openpyxl as xl

global sheet
path = r'C:\Users\ASUS\Documents' # This varaible must be changed to where you want the library Catalog saved!
files = listdir(path)
sheet = 'Library_Users.xlsx' # A blank .xlsx must be created in the specified file path

wb = xl.load_workbook(os.path.join(path, sheet))

ws = wb['Sheet1']
ws.create_sheet('Sheet2')
booklist = wb['Sheet2']
wb.save(sheet)

ws['A1'].value = 'Name'
ws['B1'].value = 'Email'
ws['C1'].value = 'Age'
ws['D1'].value = 'Password'
ws['E1'].value = 'User ID'
ws['F1'].value = 'Books'
ws['G1'].value = 'Volunteer Hours'

booklist['A1'].value = 'Book'
booklist['B1'].value = 'Author'
booklist['C1'].value = 'Rented'

class user:
    
    global count
    count = [ws.cell(row=i, column=1).value for i in range(1,ws.max_columnumn+1)]
    
    def __init__(self, name, email, age, password, userID):
        self.name = name
        self.userID = userID
        self.email = email
        self.age = age
        self.password = password
        self.books = []
        x = len(count)
        y = 1
        ws.cell(row=x, column=y, value = name)
        wb.save(sheet)
        y += 1
        ws.cell(row=x, column=y, value = email)
        wb.save(sheet)
        y += 1
        ws.cell(row=x, column=y, value = age)
        wb.save(sheet)
        y += 1
        ws.cell(row=x, column=y, value = password)
        wb.save(sheet)
        y += 1
        ws.cell(row=x, column=y, value = userID)
        wb.save(sheet)
        y += 1
        count.append(y)
        
    def __str__(self):
        return f'{self.name} \n {self.email} \n {self.age} \n {self.password}\n {self.books}'
    
    def print_name(self):
        print(self.name)
    
    def print_email(self):
        print(self.email)
    
    def print_age(self): # Returns an integer value
        print(self.age)
    
    def print_password(self):
        print(self.password)
    
    def print_userID(self): # Will return boolean value
        print(self.userID)
    
    
    def num_of_users(self): # Will return an integer value
        return len(count)
    
    def sign_in(username, password, email):
        users = [ws.cell(row=i,column=1).value for i in range(1,ws.max_columnumn+1)]
        userid = [ws.cell(row=i,column=5).value for i in range(1,ws.max_columnumn+1)]
        passwords = [ws.cell(row=i,column=4).value for i in range(1,ws.max_columnumn+1)]
        emails = [ws.cell(row=i,column=2).value for i in range(1,ws.max_columnumn+1)]
        if (username in users or username in userid) and password in passwords and email in emails:
            return True
        else:
            return False

    def rent_book(index):
        print("What book would you like to take out?: ")
        book = str.title(input(""))
        exist = False
        for i in range(len(booklist.max_row())):
            if book == booklist.cell(row=i, cloumn=1).value() and ws.cell(row = index, column = 5).value().contains(book) != True:
                exist = True
                break
        if exist:
            book += ws.cell(row = 5, cloumn = index).value + " " + book
            ws.cell(row = 5, column = index, value = book)
            print("Successfully taken out " + book + "\n")
            wb.save(sheet)
        else:
            print("Sadly, " + book + " is not in our current catalogue, but we are always looking for more books to add!\n")
    
    def report_hours(index):
        hours = float(input("Please enter the number of hours you have worked since last update \n An hour and 30 minutes should be input as 1.5 hours\n"))
        hours += ws.cell(row = index, column = 7).value 
        ws.cell(row = index, column = 5, value = hours)
        wb.save(sheet)
    
    def catalouge_update():
        book_title = str.title(input("Please enter the title of the book you wish to enter\n"))
        book_author = str.title(input("Please enter the author of the book\nIf you don't know the name of the author, please put 'Author Not Listed'\n"))
        booklist.cell(row = booklist.max_row + 2, column = 1, value = book_title)
        booklist.cell(row = booklist.max_row + 2, column = 2, value = book_author)
        wb.save(sheet)
        print(book_title + ", by " + book_author + ", has been successfully added into the catalouge")
    
    def update_name(index):
        while True:
            nn = str.strip(str.title(input("Please enter your new name: ")))
            Name = re.search(np, nn)
            if Name:
                ws.cell(row = index, column = 1, value = nn)
                wb.save(sheet)
                break
            else:
                print("Please enter a valid name")
    
    def update_email(index):
        while True:
            ne = str.strip(input("Please enter your new email: "))
            Email = re.search(ep, ne)
            if Email:
                ws.cell(row = index, column = 2, value = ne)
                wb.save(sheet)
                break
            else:
                print("Please enter a valid email address: ")
    
    def update_age(index):
        while True:
            age = int(input("Please enter your new age: "))
            a = re.search(ap, str(age))
            if a:
                ws.cell(row = index, column = 3, vaulue = age)
                wb.save(sheet)
                break
            else:
                print("Please enter a valid age")

    def update_pas(index):
        while True:
            pas = input("Please enter your new password (Same rules apply): ")
            p = re.search(pp, pas)
            if p:
                ws.cell(row = index, column = 4, value = pas)
                wb.save(sheet)
                break
            else:
                print("Please enter a valid password")
    
    def check_userID(index):
        print("User ID: " + ws.cell(row = index, column = 5).value())

    def check():
        term = str.lower(input("Are you looking for the title or the Author: "))
        if term == 'title':
            title = str.strip(str.title(input("Please enter the title of the book: ")))
            exists = False
            for i in range(booklist.max_row()):
                if title == booklist.cell(row = i, column = 1).value():
                    ba = booklist.cell(row = i, column = 2).value()
                    exists = True
            
            if exists:
                print(title + " by " + ba + " is currently in the library")
            else:
                print("We do not currently have " + title)
        else:
            author = str.strip(str.title(input("Please enter the author of the book: ")))
            exists = False
            b = []
            for i in range(booklist.max_row()):
                if author == booklist.cell(row = i, column = 2).value():
                    b.append(booklist.cell(row = i, column = 1).value())
                    exists = True
            
            if exists:
                print("We have the following books written by " + author)
                for i in range(len(b)):
                    print(b[i])
            else:
                print("We currently do not have any books written by " + author)
        
    def manage_books(index):
        if ws.cell(row = index, column = 6).value() >= 1:
            print("You currently you have the following books out:")
            books = ws.cell(row = index, column = 6).value()
            for i in range(len(ws.cell(row = index, column = 6).value())):
                print(books[i])
        else:
            print("You currently have no books rented out")
    
    def comp(index):
        if ws.cell(row = index, column = 7).value > 0:
            hours = ws.cell(row = index, column = 7).value()
            tax = (15.3 / 100) * hours
            rate = 25.10 * hours
            pay = rate - tax
            print("So far, you have earned $" + str(pay) + " based off of hours logged (after tax)")
        else:
            print("You currently have no hours logged")



# Seperation of User and Primary Code




users = []
special = ["!", ",", "/", "?", ".", "#", "$", "^", "&", "*", "(", ")", "{", "}", "|", ":", ";", "'", '"', "<", ">", " "]

np = r"(\b[a-zA-Z]+\b) (\b[a-zA-Z]+\b)"
ep = r"(\b[a-zA-Z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b)"
ap = r"^(1[89]|[2-9]\d|\d{3,})$"
pp = r"^(?=.*[A-Z])(?=.*[a-z])(?=.*[!@#$%^&*()_+{}[\]:;<>,.?~\\-]).{8,}$"

while True:
    account = str.strip(str.lower(input("Do you already have an account? (Y/N): ")))
    if account == 'y':
        break
    name = str.strip(input("Please enter your name : "))
    email = str.strip(input("Please enter your email address: "))
    age = int(input("Please enter your age (must be over 18 to create a new acc): "))
    print("Please enter your password")
    password = input("Must be at least 8 characters, have one uppercase letter, one lowercase letter, and one special character: \n")


    Name = re.search(np, name)
    Email = re.search(ep, email)
    Age = re.search(ap, str(age))
    Pas = re.search(pp, password)
    sp = random.choice(special) # 41 - 53 Creation of the user ID
    names = name.split(" ")
    fname = names[0]
    lname = names[1]
      
    while True: # Chooses two unique and random special characters
        sp_ = random.choice(special)
        if sp_ != sp:
            break

    num = int(random.randint(1000, 10000))
    ln = len(lname)
    UserID = sp_ + fname[0] + fname[2] + str(num % 10) + lname[ln - 1] + lname[ln - 2] + str(num) + sp
    if Name and Email and Age and Pas:
        new = user(name, email, age, password, UserID)
        users.append(new)
    elif Name and Email and Age and Pas != True:
        print("Invalid Password")
    elif Name and Age and Pas and Email != True:
        print( "Invalud Email")
    elif Name and Pas and Email and Age != True:
        print("Invalid Age")
    else:
        print("Invalid Name")

while True:
    print(user.print_userID())
    username = str.strip(input("Please enter your username (name or userID): "))
    email = str.strip(input("Please enter your email: "))
    password = input("Please enter your password: ")
    if user.sign_in(username, password, email):
        if username in [ws.cell(row=i,column=5).value for i in range(1,ws.max_columnumn+1)]:
            name = username
            v = 1
            while name == username:
                v += 1
                if username == ws.cell(row=v, column=1).value:
                    name = ws.cell(row=1, column=v).value
            break
        else:
            if username in [ws.cell(row=i,column=5).value for i in range(1,ws.max_columnumn+1)]:
                name = username
                v = 1
                while name == username:
                    v += 1
                    if username == ws.cell(row=v, column=5).value:
                        name = ws.cell(row=v, column=5).value
            break
    else:
        print("Incorrect Username / Password / Email\n")

while True:
    print("What would you like to do now?")
    print("1 - Rent/Return a book")
    print("2 - Report Volunteer/Work Hours")
    print("3 - Check Catalog")
    print("4 - Catalog Update")
    print("5 - Acc. Management")
    print("6 - Marketing")
    print("7 - Manage your loans")
    print("8 - Compensation Calculation")
    print("9 - Quit (Closing the program outside of this method may cause data corruption)")
    choice = int(input())
    if choice == 1:
        user.rent_book(v)
    if choice == 2:
        user.report_hours(v)
    if choice == 3:
        user.catalouge_update()
    if choice == 4:
        user.check()
    if choice == 5:
        print("What would you like to check/update?")
        print("1 - Update Name")
        print("2 - Update Email")
        print("3 - Update Age")
        print("4 - Update Password")
        print("5 - Check User ID")
        val = int(input())
        if val == 1:
            user.update_name(v)
        if val == 2:
            user.update_email(v)
        if val == 3:
            user.update_age(v)
        if val == 4:
            user.update_pas(v)
        if val == 5:
            user.check_userID(v)
    if choice == 6:
        user.manage_books(v)
    if choice == 7:
        user.comp(v)
    if choice == 8:
        print("Are you sure (y/n)")
        fn = str.lower(input())
        if fn == 'y':
            quit()